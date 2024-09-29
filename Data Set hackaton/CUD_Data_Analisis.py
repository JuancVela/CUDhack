# -*- coding: utf-8 -*-
"""
Created on Fri Sep 27 18:20:38 2024

@author: santiago.ruiz
"""
    
# In[1]:### Libraries

import pandas as pd
import os
import openpyxl
import re
    
# In[2]:### Read Data Frames

path = os.getcwd() 
folder = path + "\Data Base"

#Read Bioreactor files
folder_Bioreactor= folder+"\Bioractor"
excel_files = [file for file in os.listdir(folder_Bioreactor)]

Bioreactor_dfs = [] 

for file_name in excel_files:
    file_path = os.path.join(folder_Bioreactor, file_name)
    wb = openpyxl.load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet.sheet_state == 'visible':
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            match = re.search(r'\d+', file_name)
            if match:
                id_file = str(match.group(0))
                df['ID'] = id_file
                df.columns = [col.replace(id_file+"_", "") for col in df.columns]
            Bioreactor_dfs.append(df)
            

Bioreactor = pd.concat(Bioreactor_dfs, axis=0, ignore_index=True)

#Read Centrifuga files

folder_Centrifuga = folder+"\Centrifuga"
excel_files = [file for file in os.listdir(folder_Centrifuga)]
Centrifuga_dfs = []  # Store DataFrames of visible sheets

for file_name in excel_files:
    file_path = os.path.join(folder_Centrifuga, file_name)

    # Open the workbook using openpyxl
    wb = openpyxl.load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Check sheet visibility 
        if sheet.sheet_state == 'visible':
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            match = re.search(r'\d+', file_name)
            if match:
                id_file = str(match.group(0))
                df['ID'] = id_file
                df.columns = [col.replace(id_file+"_", "") for col in df.columns]
            Centrifuga_dfs.append(df)

Centrifuga = pd.concat(Centrifuga_dfs, axis=0, ignore_index=True)

#Read the rest of the Files on DB 
excel_files = [file for file in os.listdir(folder) if ".xlsx" in file]

#Filter only excel files 
DB = {}  # Initialize an empty dictionary to store DataFrames

for file_name in excel_files:
    df = pd.read_excel(folder + "\\" + file_name, sheet_name=None)
    for sheet_name in df.keys():
        df_name = f"{file_name}_{sheet_name}" 
        DB[df_name] = pd.read_excel(folder + "\\" + file_name, sheet_name=sheet_name) 

print(DB.keys())

#Change Key Names

key_mappings = {
    'Cinéticos IPC.xlsx_Inóculos': 'C_Inoculo',
    'Cinéticos IPC.xlsx_Cultivos finales': 'C_finales',
    'Cinéticos IPC.xlsx_Centrifugación': 'C_Centrifuga',
    'Fases producción.xlsx_Preinóculo': 'FP_Preinoculo',
    'Fases producción.xlsx_Inóculo': 'FP_Inoculo',
    'Fases producción.xlsx_Cultivo final': 'FP_final',
    'Fases producción_test.xlsx_Cultivo final': 'FP_test_Cf',
    'Horas inicio fin centrífugas.xlsx_Hoja1': 'Hrs_Centrifuga',
    'Movimientos componentes.xlsx_Full1': 'Mov_comp',
    'OF 123456.xlsx_Sheet1': 'OF',
    'Temperaturas y humedades.xlsx_WData': 'TyH_WDato',
    'Temperaturas y humedades.xlsx_Datos': 'TyH_Datos'
    # Add more mappings for other keys you want to change 
}

# 2. Update the dictionary using the mappings
for old_key, new_key in key_mappings.items():
    if old_key in DB:
        DB[new_key] = DB.pop(old_key)
        
# In[3]:### Transform Data Frames



